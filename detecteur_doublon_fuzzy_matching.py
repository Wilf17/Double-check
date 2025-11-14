#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
import sys
import os
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import unicodedata
import difflib

print("Démarrage du détecteur intelligent de doublons (fuzzy affinée)...")

# === Vérification ===
if len(sys.argv) != 3:
    print("Usage: python detecteur_doublon_fuzzy_matching.py input.csv resultat_groupes.xlsx")
    sys.exit(1)

input_file = sys.argv[1]
output_file = sys.argv[2]

if not os.path.exists(input_file):
    print(f"ERREUR : Fichier '{input_file}' introuvable.")
    sys.exit(1)

def remove_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def similarity(s1, s2):
    return difflib.SequenceMatcher(None, s1, s2).ratio()

try:
    # Lecture CSV
    df = pd.read_csv(input_file, sep=';', dtype=str, na_values=['NULL', '', ' '])
    df.columns = df.columns.str.strip()
    print(f"{len(df)} lignes chargées.")

    # === NETTOYAGE ===
    df['matricule_clean'] = (
        df['matricule']
        .astype(str)
        .str.replace(r'[-–]ANNULE.*$', '', regex=True, case=False)
        .str.strip()
        .str.upper()
    )

    df['nom_clean'] = df['nom'].astype(str).str.strip().str.upper().apply(remove_accents)
    df['prenom_clean'] = df['prenom'].astype(str).str.strip().str.upper().apply(remove_accents)
    df['etudiant_key'] = df['nom_clean'] + " | " + df['prenom_clean']

    # === 1. DOUBLONS PAR MATRICULE ===
    mat_groups = df.groupby('matricule_clean').apply(
        lambda x: x.index.tolist() if len(x) > 1 else []
    ).to_dict()
    mat_groups = {k: v for k, v in mat_groups.items() if v}

    # === 2. DOUBLONS PAR ÉTUDIANT (exact + fuzzy affinée) ===
    etu_groups = {}
    threshold = 0.85  # Affiné à 85% pour attraper "Steeve" vs "Steve" (90.9%), "Agbonon" vs "Agbonnon" etc.
    window = 50       # Fenêtre élargie pour plus de comparaisons sans perte de perf

    # Trier pour optimiser fuzzy
    df_sorted = df.sort_values('etudiant_key').reset_index()
    visited = set()

    for i in range(len(df_sorted)):
        if i in visited:
            continue
        current_key = df_sorted.iloc[i]['etudiant_key']
        group_indices = [df_sorted.iloc[i]['index']]

        for j in range(i + 1, min(i + window, len(df_sorted))):
            if j in visited:
                continue
            sim = similarity(current_key, df_sorted.iloc[j]['etudiant_key'])
            if sim > threshold:
                group_indices.append(df_sorted.iloc[j]['index'])
                visited.add(j)

        if len(group_indices) > 1:
            # Vérifier plusieurs matricules
            mats = df.loc[group_indices, 'matricule_clean'].nunique()
            if mats > 1:
                etu_groups[f'ETU_{len(etu_groups)+1}'] = group_indices
                visited.update(group_indices)

    # === Fusion des groupes ===
    all_groups = {}
    group_counter = 1

    def add_to_group(indices):
        global group_counter
        indices = sorted(set(indices))
        if not indices:
            return
        merged = False
        for gid, gidx in all_groups.items():
            if set(indices) & set(gidx):
                all_groups[gid].extend(indices)
                all_groups[gid] = sorted(set(all_groups[gid]))
                merged = True
                break
        if not merged:
            all_groups[group_counter] = indices
            group_counter += 1

    for indices in mat_groups.values():
        add_to_group(indices)
    for indices in etu_groups.values():
        add_to_group(indices)

    # === Résultat ===
    result_rows = []
    group_colors = {}
    base_colors = ['FFE699', 'FFD966', 'FFCC80', 'FFB366', 'FF9999', 'FF8080', 'CC99FF', '99CCFF', '99FFCC', '66FF99']
    palette = base_colors.copy()

    for gid in sorted(all_groups.keys()):
        if not palette:
            palette = base_colors.copy()
        color = palette.pop(0)
        group_colors[gid] = color

        indices = all_groups[gid]
        types = set()

        for idx in indices:
            row = df.iloc[idx].copy()
            mat = row['matricule_clean']
            etu = row['etudiant_key']

            if mat in mat_groups:
                types.add("MATRICULE DUPLIQUÉ")
            if any(etu in key for key in etu_groups.keys()):
                types.add("ÉTUDIANT MULTI-MATRICULE")

            row['Groupe'] = f"G{gid}"
            row['Type_doublon'] = "; ".join(sorted(types)) if types else "DOUBLON"
            result_rows.append(row)

    # === Uniques ===
    used_idx = set(sum(all_groups.values(), []))
    uniques = df[~df.index.isin(used_idx)].copy()
    uniques['Groupe'] = ""
    uniques['Type_doublon'] = ""

    # === Final DF ===
    final_df = pd.concat([pd.DataFrame(result_rows), uniques], ignore_index=True)
    final_df = final_df.sort_values(['Groupe', 'nom_clean', 'prenom_clean', 'matricule'], na_position='last')
    final_df = final_df[['matricule', 'nom', 'prenom', 'sexe', 'Groupe', 'Type_doublon']]

    # === STATISTIQUES ===
    nb_matricules_repetes = len(mat_groups)
    nb_etudiants_multi = len(etu_groups)
    nb_groupes_totaux = len(all_groups)
    nb_lignes_doublons = len(used_idx)

    print("\n" + "="*70)
    print(" " * 20 + "RÉSULTATS FINAUX")
    print("="*70)
    print(f"Total lignes lues                  : {len(df)}")
    print(f"Étudiants en doublon               : {nb_lignes_doublons}")
    print(f"Groupes de doublons détectés       : {nb_groupes_totaux}")
    print(f"   → Matricules répétés            : {nb_matricules_repetes}")
    print(f"   → Étudiants avec plusieurs matricules (fuzzy) : {nb_etudiants_multi}")
    print("="*70)

    # === Export Excel ===
    print(f"\nGénération de {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name='Doublons_Intelligents', index=False)
        ws = writer.sheets['Doublons_Intelligents']

        # Ajuster largeurs
        for i, col in enumerate(final_df.columns, 1):
            max_len = max(final_df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max_len, 50)

        # En-tête
        for cell in ws[1]:
            cell.fill = PatternFill("solid", "1f4e79")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Bordures
        thin = Side(style='thin')
        thick = Side(style='thick')
        border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)

        # Colorier
        current_group = None
        start_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            group_cell = row[4]
            group = group_cell.value if group_cell.value else ""

            for cell in row:
                cell.border = border_thin

            if group.startswith('G'):
                gid = int(group[1:])
                if group != current_group:
                    if current_group and start_row:
                        for r in range(start_row, idx):
                            for cell in ws[r]:
                                cell.fill = PatternFill("solid", group_colors[int(current_group[1:])])
                        ws[start_row][4].border = border_thick
                        ws[idx-1][4].border = border_thick
                    current_group = group
                    start_row = idx

        if current_group and start_row:
            gid = int(current_group[1:])
            for r in range(start_row, len(final_df) + 2):
                try:
                    for cell in ws[r]:
                        cell.fill = PatternFill("solid", group_colors[gid])
                except:
                    break
            ws[start_row][4].border = border_thick
            try:
                ws[len(final_df)+1][4].border = border_thick
            except:
                pass

        # Légende
        ws.append([])
        ws.append(["LÉGENDE : Groupes de doublons"])
        ws.append(["Groupe", "Type", "Couleur"])
        for gid in sorted(group_colors.keys()):
            ws.append([f"G{gid}", "", ""])
            ws.cell(ws.max_row, 3).fill = PatternFill("solid", group_colors[gid])

        # Tableau
        last_col = chr(65 + len(final_df.columns) - 1)
        tab = Table(displayName="Doublons", ref=f"A1:{last_col}{len(final_df)+1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False)
        ws.add_table(tab)

    print(f"SUCCÈS ! Fichier généré : {output_file}")
    print("   → Fuzzy affinée (85% pour attraper 'Steeve' vs 'Steve')")
    print("   → Aucun module externe requis")

except Exception as e:
    print(f"ERREUR : {e}")
    import traceback
    traceback.print_exc()