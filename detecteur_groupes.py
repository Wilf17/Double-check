#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
import sys
import os
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from unidecode import unidecode 

print("Démarrage du détecteur intelligent de doublons (insensible casse/accents)...")

# === Vérification ===
if len(sys.argv) != 3:
    print("Usage: python detecteur_doublons_intelligent_FINAL.py input.csv resultat_groupes.xlsx")
    sys.exit(1)

input_file = sys.argv[1]
output_file = sys.argv[2]

if not os.path.exists(input_file):
    print(f"ERREUR : Fichier '{input_file}' introuvable.")
    sys.exit(1)

try:
    # Lecture CSV
    df = pd.read_csv(input_file, sep=';', dtype=str, na_values=['NULL', '', ' '])
    df.columns = df.columns.str.strip()
    print(f"{len(df)} lignes chargées.")

    # === NETTOYAGE INTELLIGENT ===
    # Matricule : enlever -ANNULE, --ANNULE, espaces, casse
    df['matricule_clean'] = (
        df['matricule']
        .astype(str)
        .str.replace(r'[-–]ANNULE.*$', '', regex=True, case=False)
        .str.strip()
        .str.upper()
    )

    # Nom & Prénom : majuscules + suppression accents
    df['nom_clean'] = df['nom'].astype(str).str.strip().str.upper().apply(unidecode)
    df['prenom_clean'] = df['prenom'].astype(str).str.strip().str.upper().apply(unidecode)

    # Clé étudiant : NOM | PRENOM (insensible casse/accents)
    df['etudiant_key'] = df['nom_clean'] + " | " + df['prenom_clean']

    # === 1. DOUBLONS PAR MATRICULE ===
    mat_doublons = df.groupby('matricule_clean').filter(lambda x: len(x) > 1)
    mat_groups = {}
    if not mat_doublons.empty:
        mat_groups = mat_doublons.groupby('matricule_clean').apply(lambda x: x.index.tolist()).to_dict()

    # === 2. ÉTUDIANT AVEC PLUSIEURS MATRICULES ===
    etu_multi = df.groupby('etudiant_key').filter(lambda x: x['matricule_clean'].nunique() > 1)
    etu_groups = {}
    if not etu_multi.empty:
        etu_groups = etu_multi.groupby('etudiant_key').apply(lambda x: x.index.tolist()).to_dict()

    # === Fusion des groupes ===
    all_groups = {}
    group_counter = [1]  # Défini avant (liste mutable pour permettre l'incrémentation depuis la fonction imbriquée)

    def add_to_group(indices):
        indices = sorted(set(indices))
        if not indices:
            return
        merged = False
        for gid, gidx in all_groups.items():
            if set(indices) & set(gidx):
                all_groups[gid].extend(indices)
                all_groups[gid] = sorted(set(all_groups[gid]))
                merged = True
                break
        if not merged:
            all_groups[group_counter[0]] = indices
            group_counter[0] += 1

    for indices in mat_groups.values():
        add_to_group(indices)
    for indices in etu_groups.values():
        add_to_group(indices)

    # === Résultat ===
    result_rows = []
    group_colors = {}
    base_colors = ['FFE699', 'FFD966', 'FFCC80', 'FFB366', 'FF9999', 'FF8080', 'CC99FF', '99CCFF', '99FFCC', '66FF99']
    palette = base_colors.copy()

    for gid in sorted(all_groups.keys()):
        if not palette:
            palette = base_colors.copy()
        color = palette.pop(0)
        group_colors[gid] = color

        indices = all_groups[gid]
        types = set()

        for idx in indices:
            row = df.iloc[idx].copy()
            mat = row['matricule_clean']
            etu = row['etudiant_key']

            if mat in mat_groups and any(idx in indices_list for indices_list in mat_groups.values()):
                types.add("MATRICULE DUPLIQUÉ")
            if etu in etu_groups and any(idx in indices_list for indices_list in etu_groups.values()):
                types.add("ÉTUDIANT MULTI-MATRICULE")

            row['Groupe'] = f"G{gid}"
            row['Type_doublon'] = "; ".join(sorted(types)) if types else "DOUBLON"
            result_rows.append(row)

    # === Uniques ===
    used_idx = set(sum(all_groups.values(), []))
    uniques = df[~df.index.isin(used_idx)].copy()
    uniques['Groupe'] = ""
    uniques['Type_doublon'] = ""

    # === Final DF ===
    final_df = pd.concat([pd.DataFrame(result_rows), uniques], ignore_index=True)
    final_df = final_df.sort_values(['Groupe', 'nom_clean', 'prenom_clean', 'matricule'], na_position='last')
    final_df = final_df[['matricule', 'nom', 'prenom', 'sexe', 'Groupe', 'Type_doublon']]  # Colonnes d'origine

    # === STATISTIQUES FINALES ===
    nb_matricules_repetes = len(mat_groups)
    nb_etudiants_multi_mat = len(etu_groups)
    nb_groupes_totaux = len(all_groups)
    nb_lignes_doublons = len(used_idx)

    print("\n" + "="*70)
    print(" " * 20 + "RÉSULTATS FINAUX")
    print("="*70)
    print(f"Total lignes lues                  : {len(df)}")
    print(f"Étudiants en doublon               : {nb_lignes_doublons}")
    print(f"Groupes de doublons détectés       : {nb_groupes_totaux}")
    print(f"   → Matricules répétés            : {nb_matricules_repetes}")
    print(f"   → Étudiants avec plusieurs matricules : {nb_etudiants_multi_mat}")
    print("="*70)

    # === Export Excel ===
    print(f"\nGénération de {output_file}...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name='Doublons_Intelligents', index=False)
        ws = writer.sheets['Doublons_Intelligents']

        # Ajuster largeurs
        for i, col in enumerate(final_df.columns, 1):
            max_len = max(final_df[col].astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max_len, 50)

        # En-tête
        for cell in ws[1]:
            cell.fill = PatternFill("solid", "1f4e79")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Bordures
        thin = Side(style='thin')
        thick = Side(style='thick')
        border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
        border_thick = Border(top=thick, left=thick, right=thick, bottom=thick)

        # Colorier les groupes
        current_group = None
        start_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            group_cell = row[4]
            group = group_cell.value if group_cell.value else ""

            for cell in row:
                cell.border = border_thin

            if group.startswith('G'):
                gid = int(group[1:])
                if group != current_group:
                    if current_group and start_row:
                        for r in range(start_row, idx):
                            for cell in ws[r]:
                                cell.fill = PatternFill("solid", group_colors[int(current_group[1:])])
                        ws[start_row][4].border = border_thick
                        ws[idx-1][4].border = border_thick
                    current_group = group
                    start_row = idx

        if current_group and start_row:
            gid = int(current_group[1:])
            for r in range(start_row, len(final_df) + 2):
                try:
                    for cell in ws[r]:
                        cell.fill = PatternFill("solid", group_colors[gid])
                except:
                    break
            ws[start_row][4].border = border_thick
            try:
                ws[len(final_df)+1][4].border = border_thick
            except:
                pass

        # Légende
        ws.append([])
        ws.append(["LÉGENDE : Groupes de doublons"])
        ws.append(["Groupe", "Type", "Couleur"])
        for gid in sorted(group_colors.keys()):
            ws.append([f"G{gid}", "", ""])
            ws.cell(ws.max_row, 3).fill = PatternFill("solid", group_colors[gid])

        # Tableau
        last_col = chr(65 + len(final_df.columns) - 1)
        tab = Table(displayName="Doublons", ref=f"A1:{last_col}{len(final_df)+1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=False)
        ws.add_table(tab)

    print(f"SUCCÈS ! Fichier généré : {output_file}")
    print("   → Insensible à la casse et aux accents")
    print("   → Matricules -ANNULE ignorés")
    print("   → Statistiques complètes affichées")

except Exception as e:
    print(f"ERREUR : {e}")
    import traceback
    traceback.print_exc()